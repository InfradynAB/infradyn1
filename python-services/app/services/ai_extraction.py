"""
AI Extraction Service
Handles OCR (AWS Textract) and GPT parsing for document extraction
"""
import io
import re
import json
import asyncio
from typing import Optional, Dict, Any, List
from pathlib import Path

import boto3
import httpx
from openai import OpenAI
import pdfplumber
from docx import Document as DocxDocument
import openpyxl

from app.config import get_settings


class AIExtractionService:
    """
    AI-powered document extraction service
    
    Features:
    - AWS Textract for OCR (PDF, images)
    - pdfplumber for direct PDF text extraction
    - python-docx for Word documents
    - openpyxl for Excel files
    - OpenAI GPT-4 for structured data parsing
    """
    
    def __init__(self):
        settings = get_settings()
        
        # AWS Textract client
        self.textract = boto3.client(
            "textract",
            region_name=settings.aws_region,
            aws_access_key_id=settings.aws_access_key_id,
            aws_secret_access_key=settings.aws_secret_access_key,
        )
        
        # S3 client for fetching files
        self.s3 = boto3.client(
            "s3",
            region_name=settings.aws_region,
            aws_access_key_id=settings.aws_access_key_id,
            aws_secret_access_key=settings.aws_secret_access_key,
        )
        
        # OpenAI client
        self.openai = OpenAI(api_key=settings.openai_api_key)
        
        self.s3_bucket = settings.aws_s3_bucket
    
    # =========================================================================
    # PUBLIC METHODS
    # =========================================================================
    
    async def extract_purchase_order(self, file_url: str) -> Dict[str, Any]:
        """Extract structured PO data from a file URL"""
        try:
            # Get file extension
            ext = self._get_extension(file_url)
            
            # Extract raw text based on file type
            if ext in [".xlsx", ".xls"]:
                raw_text = await self._extract_excel(file_url)
            elif ext in [".docx", ".doc"]:
                raw_text = await self._extract_word(file_url)
            elif ext == ".pdf":
                raw_text = await self._extract_pdf(file_url)
            elif ext in [".png", ".jpg", ".jpeg"]:
                raw_text = await self._extract_image(file_url)
            else:
                return {"success": False, "error": f"Unsupported file type: {ext}"}
            
            if not raw_text:
                return {"success": False, "error": "Could not extract text from document"}
            
            # Parse with GPT
            parsed_data = await self._parse_po_with_gpt(raw_text)
            parsed_data["raw_text"] = raw_text[:5000]  # Include truncated raw text
            
            return {"success": True, "data": parsed_data}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def extract_invoice(self, file_url: str) -> Dict[str, Any]:
        """Extract structured invoice data from a file URL"""
        try:
            ext = self._get_extension(file_url)
            
            if ext in [".xlsx", ".xls"]:
                raw_text = await self._extract_excel(file_url)
            elif ext in [".docx", ".doc"]:
                raw_text = await self._extract_word(file_url)
            elif ext == ".pdf":
                raw_text = await self._extract_pdf(file_url)
            elif ext in [".png", ".jpg", ".jpeg"]:
                raw_text = await self._extract_image(file_url)
            else:
                return {"success": False, "error": f"Unsupported file type: {ext}"}
            
            if not raw_text:
                return {"success": False, "error": "Could not extract text from document"}
            
            parsed_data = await self._parse_invoice_with_gpt(raw_text)
            parsed_data["raw_text"] = raw_text[:5000]
            
            return {"success": True, "data": parsed_data}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def extract_milestones(self, file_url: str) -> Dict[str, Any]:
        """Extract milestone schedule from a file"""
        try:
            ext = self._get_extension(file_url)
            
            # Excel is preferred for milestones
            if ext in [".xlsx", ".xls"]:
                milestones = await self._extract_milestones_from_excel(file_url)
            elif ext == ".pdf":
                raw_text = await self._extract_pdf(file_url)
                milestones = await self._parse_milestones_with_gpt(raw_text)
            else:
                return {"success": False, "error": f"Unsupported file type for milestones: {ext}"}
            
            return {"success": True, "data": {"milestones": milestones}}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def extract_from_bytes(
        self, 
        content: bytes, 
        filename: str, 
        document_type: str
    ) -> Dict[str, Any]:
        """Extract from raw file bytes (for uploads)"""
        try:
            ext = Path(filename).suffix.lower()
            
            # Extract text based on file type
            if ext in [".xlsx", ".xls"]:
                raw_text = self._extract_excel_from_bytes(content)
            elif ext in [".docx", ".doc"]:
                raw_text = self._extract_word_from_bytes(content)
            elif ext == ".pdf":
                raw_text = self._extract_pdf_from_bytes(content)
            else:
                return {"success": False, "error": f"Unsupported file type: {ext}"}
            
            if not raw_text:
                return {"success": False, "error": "Could not extract text from document"}
            
            # Parse based on document type
            if document_type == "purchase_order":
                parsed_data = await self._parse_po_with_gpt(raw_text)
            elif document_type == "invoice":
                parsed_data = await self._parse_invoice_with_gpt(raw_text)
            elif document_type == "milestone":
                milestones = await self._parse_milestones_with_gpt(raw_text)
                parsed_data = {"milestones": milestones}
            else:
                return {"success": False, "error": f"Unknown document type: {document_type}"}
            
            parsed_data["raw_text"] = raw_text[:5000]
            return {"success": True, "data": parsed_data}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    # =========================================================================
    # TEXT EXTRACTION METHODS
    # =========================================================================
    
    async def _extract_pdf(self, file_url: str) -> Optional[str]:
        """
        Extract text from PDF.
        Strategy:
        1. Try pdfplumber first (fast, works well for native text PDFs)
        2. If text is too short (<500 chars), fall back to Textract (better for scanned docs)
        """
        try:
            content = await self._download_file(file_url)
            
            # Try pdfplumber first
            text = self._extract_pdf_from_bytes(content)
            
            # If pdfplumber got reasonable text, use it
            if text and len(text.strip()) > 500:
                print(f"[PDF Extract] pdfplumber succeeded: {len(text)} chars")
                return text
            
            # Fallback to Textract for scanned/image PDFs
            print(f"[PDF Extract] pdfplumber got {len(text) if text else 0} chars, falling back to Textract")
            return await self._extract_pdf_with_textract(file_url)
            
        except Exception as e:
            print(f"PDF extraction error: {e}")
            return None
    
    async def _extract_pdf_with_textract(self, file_url: str) -> Optional[str]:
        """Extract text from PDF using AWS Textract async API"""
        try:
            from urllib.parse import urlparse
            parsed = urlparse(file_url)
            
            # Get bucket and key
            hostname_parts = parsed.hostname.split(".")
            bucket = hostname_parts[0]
            s3_key = parsed.path.lstrip("/")
            
            print(f"[Textract] Starting async extraction: {bucket}/{s3_key}")
            
            # Start async job
            start_response = self.textract.start_document_text_detection(
                DocumentLocation={
                    "S3Object": {
                        "Bucket": bucket,
                        "Name": s3_key
                    }
                }
            )
            
            job_id = start_response["JobId"]
            print(f"[Textract] Job started: {job_id}")
            
            # Poll for completion (max 2 minutes)
            max_attempts = 60
            for attempt in range(max_attempts):
                await asyncio.sleep(2)  # Wait 2 seconds between polls
                
                get_response = self.textract.get_document_text_detection(JobId=job_id)
                status = get_response["JobStatus"]
                
                if status == "SUCCEEDED":
                    # Collect all text blocks
                    all_blocks = get_response.get("Blocks", [])
                    
                    # Handle pagination
                    next_token = get_response.get("NextToken")
                    while next_token:
                        page_response = self.textract.get_document_text_detection(
                            JobId=job_id, NextToken=next_token
                        )
                        all_blocks.extend(page_response.get("Blocks", []))
                        next_token = page_response.get("NextToken")
                    
                    # Extract lines of text
                    lines = [
                        block["Text"]
                        for block in all_blocks
                        if block["BlockType"] == "LINE"
                    ]
                    
                    text = "\n".join(lines)
                    print(f"[Textract] Extracted {len(text)} chars from {len(all_blocks)} blocks")
                    return text
                    
                elif status == "FAILED":
                    print("[Textract] Job failed")
                    return None
                    
                if attempt % 10 == 0:
                    print(f"[Textract] Still processing... ({attempt * 2}s elapsed)")
            
            print("[Textract] Job timed out")
            return None
            
        except Exception as e:
            print(f"[Textract] Error: {e}")
            return None
    
    def _extract_pdf_from_bytes(self, content: bytes) -> Optional[str]:
        """Extract text from PDF bytes using pdfplumber"""
        try:
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text_parts = []
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
                return "\n\n".join(text_parts)
        except Exception as e:
            print(f"PDF bytes extraction error: {e}")
            return None
    
    async def _extract_word(self, file_url: str) -> Optional[str]:
        """Extract text from Word document"""
        try:
            content = await self._download_file(file_url)
            return self._extract_word_from_bytes(content)
        except Exception as e:
            print(f"Word extraction error: {e}")
            return None
    
    def _extract_word_from_bytes(self, content: bytes) -> Optional[str]:
        """Extract text from Word document bytes"""
        try:
            doc = DocxDocument(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)
        except Exception as e:
            print(f"Word bytes extraction error: {e}")
            return None
    
    async def _extract_excel(self, file_url: str) -> Optional[str]:
        """Extract text from Excel file"""
        try:
            content = await self._download_file(file_url)
            return self._extract_excel_from_bytes(content)
        except Exception as e:
            print(f"Excel extraction error: {e}")
            return None
    
    def _extract_excel_from_bytes(self, content: bytes) -> Optional[str]:
        """Extract text from Excel bytes"""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            text_parts = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                
                for row in sheet.iter_rows():
                    row_values = [str(cell.value) if cell.value else "" for cell in row]
                    if any(row_values):
                        text_parts.append(" | ".join(row_values))
            
            return "\n".join(text_parts)
        except Exception as e:
            print(f"Excel bytes extraction error: {e}")
            return None
    
    async def _extract_image(self, file_url: str) -> Optional[str]:
        """Extract text from image using AWS Textract"""
        try:
            # Parse S3 key from URL
            s3_key = self._parse_s3_key(file_url)
            
            response = self.textract.detect_document_text(
                Document={
                    "S3Object": {
                        "Bucket": self.s3_bucket,
                        "Name": s3_key
                    }
                }
            )
            
            # Extract text blocks
            lines = []
            for block in response.get("Blocks", []):
                if block["BlockType"] == "LINE":
                    lines.append(block.get("Text", ""))
            
            return "\n".join(lines)
            
        except Exception as e:
            print(f"Image extraction error: {e}")
            return None
    
    # =========================================================================
    # GPT PARSING METHODS
    # =========================================================================
    
    async def _parse_po_with_gpt(self, raw_text: str) -> Dict[str, Any]:
        """Parse raw text into structured PO data using GPT-4"""
        
        prompt = f"""Analyze this Purchase Order document and extract structured data.
Return a JSON object with these fields (use null for missing values):

{{
    "po_number": "string or null",
    "vendor_name": "string or null", 
    "date": "YYYY-MM-DD or null",
    "total_value": number or null,
    "currency": "3-letter code like USD, EUR, KES or null",
    "scope": "brief description of work scope or null",
    "payment_terms": "e.g., Net 30, 50% advance or null",
    "incoterms": "e.g., FOB, CIF, EXW or null",
    "retention_percentage": number (0-100) or null,
    "milestones": [
        {{
            "title": "string",
            "description": "string or null",
            "expected_date": "YYYY-MM-DD or null",
            "payment_percentage": number
        }}
    ],
    "boq_items": [
        {{
            "item_number": "string",
            "description": "string",
            "unit": "string",
            "quantity": number,
            "unit_price": number,
            "total_price": number
        }}
    ],
    "confidence": number between 0 and 1
}}

Document text:
{raw_text[:15000]}
"""
        
        try:
            response = self.openai.chat.completions.create(
                model="gpt-4-turbo-preview",
                messages=[
                    {"role": "system", "content": "You are a document extraction assistant. Always respond with valid JSON only. No explanations or markdown."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=4000,
            )
            
            content = response.choices[0].message.content
            # Clean markdown code blocks if present
            content = re.sub(r'^```json\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
            
            return json.loads(content)
            
        except Exception as e:
            print(f"GPT parsing error: {e}")
            return {
                "po_number": None,
                "vendor_name": None,
                "date": None,
                "total_value": None,
                "currency": None,
                "milestones": [],
                "boq_items": [],
                "confidence": 0.0
            }
    
    async def _parse_invoice_with_gpt(self, raw_text: str) -> Dict[str, Any]:
        """Parse raw text into structured invoice data"""
        
        prompt = f"""Analyze this Invoice document and extract structured data.
Return a JSON object with these fields (use null for missing values):

{{
    "invoice_number": "string or null",
    "vendor_name": "string or null",
    "date": "YYYY-MM-DD or null",
    "due_date": "YYYY-MM-DD or null",
    "total_amount": number or null,
    "currency": "3-letter code or null",
    "subtotal": number or null,
    "tax_amount": number or null,
    "line_items": [
        {{
            "description": "string",
            "quantity": number or null,
            "unit_price": number or null,
            "amount": number
        }}
    ],
    "confidence": number between 0 and 1
}}

Document text:
{raw_text[:15000]}
"""
        
        try:
            response = self.openai.chat.completions.create(
                model="gpt-4-turbo-preview",
                messages=[
                    {"role": "system", "content": "You are a document extraction assistant. Always respond with valid JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=4000,
            )
            
            content = response.choices[0].message.content
            content = re.sub(r'^```json\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
            
            return json.loads(content)
            
        except Exception as e:
            print(f"GPT invoice parsing error: {e}")
            return {
                "invoice_number": None,
                "vendor_name": None,
                "total_amount": None,
                "confidence": 0.0
            }
    
    async def _parse_milestones_with_gpt(self, raw_text: str) -> List[Dict[str, Any]]:
        """Parse raw text into milestone list"""
        
        prompt = f"""Extract payment milestones from this document.
Return a JSON array of milestones:

[
    {{
        "title": "string",
        "description": "string or null",
        "expected_date": "YYYY-MM-DD or null",
        "payment_percentage": number (should sum to 100)
    }}
]

Document text:
{raw_text[:15000]}
"""
        
        try:
            response = self.openai.chat.completions.create(
                model="gpt-4-turbo-preview",
                messages=[
                    {"role": "system", "content": "You are a document extraction assistant. Always respond with valid JSON array only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=2000,
            )
            
            content = response.choices[0].message.content
            content = re.sub(r'^```json\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
            
            return json.loads(content)
            
        except Exception as e:
            print(f"GPT milestone parsing error: {e}")
            return []
    
    async def _extract_milestones_from_excel(self, file_url: str) -> List[Dict[str, Any]]:
        """
        Extract milestones directly from Excel structure
        Looks for common patterns in milestone schedules
        """
        try:
            content = await self._download_file(file_url)
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            
            milestones = []
            sheet = wb.active
            
            # Find header row
            header_row = None
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10)):
                values = [str(cell.value).lower() if cell.value else "" for cell in row]
                if any("milestone" in v or "payment" in v or "description" in v for v in values):
                    header_row = row_idx + 1
                    break
            
            if not header_row:
                # Fall back to GPT parsing
                raw_text = self._extract_excel_from_bytes(content)
                return await self._parse_milestones_with_gpt(raw_text)
            
            # Parse data rows
            for row in sheet.iter_rows(min_row=header_row + 1):
                values = [cell.value for cell in row]
                
                # Skip empty rows
                if not any(values):
                    continue
                
                # Try to extract milestone data (flexible column mapping)
                milestone = {}
                for idx, val in enumerate(values):
                    if val is None:
                        continue
                    val_str = str(val).strip()
                    
                    # Title (first text column)
                    if not milestone.get("title") and isinstance(val, str) and len(val) > 3:
                        milestone["title"] = val_str
                    # Percentage (number between 0-100)
                    elif isinstance(val, (int, float)) and 0 < val <= 100:
                        if "payment_percentage" not in milestone:
                            milestone["payment_percentage"] = float(val)
                    # Date
                    elif hasattr(val, "strftime"):
                        if "expected_date" not in milestone:
                            milestone["expected_date"] = val.strftime("%Y-%m-%d")
                
                if milestone.get("title") and milestone.get("payment_percentage"):
                    milestones.append({
                        "title": milestone.get("title", ""),
                        "description": None,
                        "expected_date": milestone.get("expected_date"),
                        "payment_percentage": milestone.get("payment_percentage", 0)
                    })
            
            return milestones
            
        except Exception as e:
            print(f"Excel milestone extraction error: {e}")
            return []
    
    # =========================================================================
    # HELPER METHODS
    # =========================================================================
    
    async def _download_file(self, file_url: str) -> bytes:
        """Download file from URL (S3 or direct HTTP)"""
        # Check if it's an S3 URL (either s3:// or https://*.s3.*.amazonaws.com)
        if file_url.startswith("s3://"):
            # Parse S3 URL: s3://bucket/key
            s3_key = self._parse_s3_key(file_url)
            response = self.s3.get_object(Bucket=self.s3_bucket, Key=s3_key)
            return response["Body"].read()
        elif "s3." in file_url and "amazonaws.com" in file_url:
            # HTTPS S3 URL: https://bucket.s3.region.amazonaws.com/key
            # Extract bucket and key from URL
            from urllib.parse import urlparse
            parsed = urlparse(file_url)
            
            # Get bucket from hostname (bucket.s3.region.amazonaws.com)
            hostname_parts = parsed.hostname.split(".")
            bucket = hostname_parts[0]
            
            # Key is the path without leading slash
            s3_key = parsed.path.lstrip("/")
            
            print(f"[S3 Download] Bucket: {bucket}, Key: {s3_key}")
            
            response = self.s3.get_object(Bucket=bucket, Key=s3_key)
            return response["Body"].read()
        else:
            # Regular HTTP download (for presigned URLs or external files)
            async with httpx.AsyncClient() as client:
                response = await client.get(file_url)
                response.raise_for_status()
                return response.content
    
    def _get_extension(self, file_url: str) -> str:
        """Get file extension from URL"""
        # Remove query string
        clean_url = file_url.split("?")[0]
        return Path(clean_url).suffix.lower()
    
    def _parse_s3_key(self, file_url: str) -> str:
        """Parse S3 key from various URL formats"""
        if file_url.startswith("s3://"):
            # s3://bucket/key format
            parts = file_url[5:].split("/", 1)
            return parts[1] if len(parts) > 1 else ""
        elif "s3.amazonaws.com" in file_url:
            # https://bucket.s3.amazonaws.com/key format
            from urllib.parse import urlparse
            parsed = urlparse(file_url)
            return parsed.path.lstrip("/")
        else:
            # Assume it's already a key
            return file_url
